"""
BSD 3-Clause License

Copyright (c) 2020, vj-sarthy
All rights reserved.

@Author: Vijay Sarthy, Ph.D., Mechanical Engineering, Clemson University, USA
@Email: vsreedh@g.clemson.edu


Redistribution and use in source and binary forms, with or without
modification, are permitted provided that the following conditions are met:

1. Redistributions of source code must retain the above copyright notice, this
   list of conditions and the following disclaimer.

2. Redistributions in binary form must reproduce the above copyright notice,
   this list of conditions and the following disclaimer in the documentation
   and/or other materials provided with the distribution.

3. Neither the name of the copyright holder nor the names of its
   contributors may be used to endorse or promote products derived from
   this software without specific prior written permission.

THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE
FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR
SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.'


Python code to determine the Wbal (GET) from a workout file that has power and time data.

Preparing the data, Please make sure that:
1. Time is on column 1 and is expressed in seconds.
2. Power data is on column 2 and is expressed in Watts.
3. You will need to input your CP (similar to FTP) and W' (similar to FRC) to run the code. For the example file, CP = 235 W and wprime = 10000 J

Discalimer: This is an algorithm implementing Skiba and colleagues'[1] bi-conditional exponential model of recovery of W'.
The W'bal calculated by the model may not be accurate. (Refer criticisms of Skiba models available in the literature)

"""

import os
import click
import pandas as pd
import glob
import numpy as np
from tabulate import tabulate
import matplotlib.pyplot as plt
import math
from openpyxl import *
import tempfile

class WBAL(object):
    
    def __init__(self):
        self.wbal = None
            
    def parse(self, file):
        st = pd.read_excel(file, "Sheet1")
        self.time = st['t'].values.tolist()
        self.power = st['Power'].values.tolist()
                
    def userinput(self):
        #get user inputs for cp and W'
        self.cp = int(input("Please enter the CP: "))
        self.wprime = int(input("please enter the Wprime: "))

    
    def compute(self):
        '''
        1. Check if power data point is above or below CP
        2. If the data point is greater than CP, then use the hyperbolic CP model, i.e. P = CP+(W'/t)
        3. If the data point is less than CP, then use the exponential Skiba2 model (Refer to read me file), i.e. W'bal = W'- W'expended * e^(-DCP*t/W')
        
        '''                
        # Converting lists into numerical arrays
        time = np.array(self.time)
        power  = np.array(self.power)
        
        # Assigning new variables for ease of use in for loops
        cp = self.cp
        wprime = self.wprime
        
        k = len(time)
        self.wbal = np.zeros(k)
        delt = np.zeros(k)
        delp = np.zeros(k)
        delt[0] = time[0]
        self.wbal[0] = wprime
                
        for i in range(k):
            if power[i]>cp:
                delp[i] = power[i]-cp
            else:
                delp[i] = 0
     
        for i in range(1,k):
            delt[i] = time[i]-time[i-1]
            if power[i]>=cp:
                self.wbal[i] = self.wbal[i-1]-(delp[i]*delt[i])
            elif power[i]<cp and self.wbal[i-1]>=wprime:
                self.wbal[i] = wprime
            elif power[i]<cp and self.wbal[i-1]<wprime:
                self.wbal[i] = wprime-((wprime-self.wbal[i-1])*math.exp(-1*(cp-power[i])*delt[i]/wprime))   
           
                
    def display(self):
        print("The W'bal at the end of the workout is: {} J".format(self.wbal[-1]))
                
    def plot(self):
        
        # Creating an array of length k with every entry as cp to aid in plotting
        k = len(np.array(self.time))
        cparr = np.zeros(k)
        
        for i in range(k):
            cparr[i] = self.cp
        
        fig,ax=plt.subplots()
        # Creating two lines for plottijng Power vs time and CP
        line1, = ax.plot(self.time, self.power, color ="blue", label = "Power")
        line2, = ax.plot(self.time, cparr, '--', color = "black", label = "CP")
        
        # Creating legend for Power
        first_leg = plt.legend(handles=[line1], bbox_to_anchor=(0.25, 1.1), frameon = False)
        plt.gca().add_artist(first_leg)
        
        #Creating legend for CP
        plt.legend(handles=[line2], bbox_to_anchor=(0.6, 1.1), frameon = False)
        
        ax.set_xlabel("Time (s)")
        ax.set_ylabel("Power (W)")
        ax2=ax.twinx()
        ax2.plot(self.time, self.wbal, color ="red", label = "W'bal")
        plt.legend(bbox_to_anchor=(0.9, 1.1), frameon = False)
        ax2.set_ylabel("W'bal (J)")
        plt.savefig('Wbal.jpg', bbox_inches='tight')
        
    def append(self,file):
        wb=load_workbook(file)
        ws=wb["Sheet1"]
        wcell1=ws.cell(1,3)
        wcell1.value="W'bal"
        k = len(self.time)
        wcellv = np.zeros(k)
        for i in range(2,k+2):
            ws.cell(i,3).value = self.wbal[i-2]
        wb.save(file)
        print("The excel file has been appended with the W'bal column. Please check.")
        
@click.command()
@click.argument('xlsfilepath')
def main(xlsfilepath):
    results = []
    if os.path.isfile(xlsfilepath):
        '''Compute W'bal'''
        try:
            obj = WBAL()
            obj.parse(xlsfilepath)
            obj.userinput()
            obj.compute()
            obj.display()
            results.append([xlsfilepath.split("\\")[-1], obj.wbal[-1]])
            obj.plot()
            obj.append(xlsfilepath)
        except Exception as e:
            print("Error while processing data file: {}, error: {}. Please check the excel file.".format(excelfile, e))
    else:
         # Path to folder containing excel files
        print("Searching for data files in the directory: {}".format(xlsfilepath))
       
        # set all .xls files in your folder to list
        allfiles = glob.glob(xlsfilepath + "\*.xlsx")

        # for loop to aquire all excel files in folder
        
        for excelfile in allfiles:
            '''Compute threshold'''
            try:
                print("Processing file: {}".format(excelfile))
                obj = WBAL()
                obj.parse(excelfile)
                obj.userinput()
                obj.compute()
                obj.append(xlsfilepath)
                results.append([excelfile.split("\\")[-1], obj.wbal[-1]])
            except Exception as e:
                print("Error while processing data file: {}, error: {}. Please correct the file.".format(excelfile, e))
    print(tabulate(results, headers=["File", "W'bal at the end (J)"]))

if __name__ == '__main__':
    main()



